"""
Download endpoint - Provide templates and sample data files.

GET /download-template - Download empty Excel template
GET /download-sample - Download sample dataset (Vidya Mandir)
GET /sample-data - Get sample data as JSON for auto-loading
"""

import io
import logging

from fastapi import APIRouter
from fastapi.responses import StreamingResponse, JSONResponse

logger = logging.getLogger(__name__)

router = APIRouter()

# Sample data - Vidya Mandir High School
SAMPLE_SCHOOL = {
    "school_id": 1,
    "name": "Vidya Mandir High School",
    "academic_year": "2025-2026",
    "start_time": "08:00",
    "end_time": "15:30",
    "weekdays": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"],
    "periods_per_weekday": 9,
    "saturday_periods": 4,
    "period_duration_minutes": 45,
    "prayer_enabled": True,
    "prayer_duration_minutes": 30,
    "lunch_period_index": 5,
    "lunch_after_period": 5,
    "lunch_duration_minutes": 40,
    "recess_period_indices": [3],
    "recess_after_every_n_periods": 3,
    "recess_duration_minutes": 20
}

SAMPLE_PERIODS = [
    {"period_index": 0, "start_time": "08:00", "end_time": "08:30", "duration_minutes": 30, "is_prayer": True, "is_recess": False, "is_lunch": False},
    {"period_index": 1, "start_time": "08:30", "end_time": "09:15", "duration_minutes": 45, "is_prayer": False, "is_recess": False, "is_lunch": False},
    {"period_index": 2, "start_time": "09:15", "end_time": "10:00", "duration_minutes": 45, "is_prayer": False, "is_recess": False, "is_lunch": False},
    {"period_index": 3, "start_time": "10:00", "end_time": "10:20", "duration_minutes": 20, "is_prayer": False, "is_recess": True, "is_lunch": False},
    {"period_index": 4, "start_time": "10:20", "end_time": "11:05", "duration_minutes": 45, "is_prayer": False, "is_recess": False, "is_lunch": False},
    {"period_index": 5, "start_time": "11:05", "end_time": "11:50", "duration_minutes": 45, "is_prayer": False, "is_recess": False, "is_lunch": False},
    {"period_index": 6, "start_time": "11:50", "end_time": "12:30", "duration_minutes": 40, "is_prayer": False, "is_recess": False, "is_lunch": True},
    {"period_index": 7, "start_time": "12:30", "end_time": "13:15", "duration_minutes": 45, "is_prayer": False, "is_recess": False, "is_lunch": False},
    {"period_index": 8, "start_time": "13:15", "end_time": "14:00", "duration_minutes": 45, "is_prayer": False, "is_recess": False, "is_lunch": False},
    {"period_index": 9, "start_time": "14:00", "end_time": "14:45", "duration_minutes": 45, "is_prayer": False, "is_recess": False, "is_lunch": False},
]

SAMPLE_TEACHERS = [
    # Mathematics Teachers (4 teachers for 10 classes)
    {"teacher_id": "T001", "name": "Rajesh Kumar", "subjects_can_teach": ["MATH"], "min_periods_day": 0, "max_periods_day": 7, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T002", "name": "Priya Sharma", "subjects_can_teach": ["MATH"], "min_periods_day": 0, "max_periods_day": 7, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T003", "name": "Amit Verma", "subjects_can_teach": ["MATH"], "min_periods_day": 0, "max_periods_day": 7, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T004", "name": "Sunita Agarwal", "subjects_can_teach": ["MATH"], "min_periods_day": 0, "max_periods_day": 7, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    
    # Science Teachers (4 teachers for 10 classes)
    {"teacher_id": "T005", "name": "Suresh Reddy", "subjects_can_teach": ["SCI", "PHY_LAB"], "min_periods_day": 0, "max_periods_day": 7, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T006", "name": "Lakshmi Nair", "subjects_can_teach": ["SCI", "PHY_LAB"], "min_periods_day": 0, "max_periods_day": 7, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T007", "name": "Raghav Menon", "subjects_can_teach": ["SCI", "CHEM_LAB"], "min_periods_day": 0, "max_periods_day": 7, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T008", "name": "Deepa Krishnan", "subjects_can_teach": ["SCI", "CHEM_LAB"], "min_periods_day": 0, "max_periods_day": 7, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    
    # English Teachers (4 teachers for 10 classes)
    {"teacher_id": "T009", "name": "Anita Desai", "subjects_can_teach": ["ENG"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 45, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T010", "name": "Meena Patel", "subjects_can_teach": ["ENG"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 45, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T011", "name": "Rahul Saxena", "subjects_can_teach": ["ENG"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 45, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T012", "name": "Neelam Gupta", "subjects_can_teach": ["ENG"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 45, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    
    # Social Studies Teachers (3 teachers for 10 classes)
    {"teacher_id": "T013", "name": "Vijay Singh", "subjects_can_teach": ["SST"], "min_periods_day": 0, "max_periods_day": 7, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T014", "name": "Geeta Rao", "subjects_can_teach": ["SST"], "min_periods_day": 0, "max_periods_day": 7, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T015", "name": "Prakash Jha", "subjects_can_teach": ["SST"], "min_periods_day": 0, "max_periods_day": 7, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    
    # Hindi Teachers (4 teachers - 2 classes each max for language block compatibility)
    {"teacher_id": "T016", "name": "Arun Kumar", "subjects_can_teach": ["HINDI"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 50, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T017", "name": "Suman Devi", "subjects_can_teach": ["HINDI"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 50, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T018", "name": "Ramesh Tiwari", "subjects_can_teach": ["HINDI"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 50, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T033", "name": "Anita Singh", "subjects_can_teach": ["HINDI"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 50, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    
    # Kannada Teachers (4 teachers - 2 classes each max for language block compatibility)
    {"teacher_id": "T019", "name": "Kavita Menon", "subjects_can_teach": ["KANNADA"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 50, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T020", "name": "Ravi Hegde", "subjects_can_teach": ["KANNADA"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 50, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T021", "name": "Shobha Shetty", "subjects_can_teach": ["KANNADA"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 50, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T034", "name": "Mohan Gowda", "subjects_can_teach": ["KANNADA"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 50, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    
    # Sanskrit Teachers (4 teachers - 2 classes each max for language block compatibility)
    {"teacher_id": "T022", "name": "Deepak Verma", "subjects_can_teach": ["SANSKRIT"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 50, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T023", "name": "Padma Sharma", "subjects_can_teach": ["SANSKRIT"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 50, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T024", "name": "Gopal Iyer", "subjects_can_teach": ["SANSKRIT"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 50, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T035", "name": "Lakshmi Pandit", "subjects_can_teach": ["SANSKRIT"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 50, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    
    # PE Teachers (3 teachers for 10 classes)
    {"teacher_id": "T025", "name": "Rohit Bhatt", "subjects_can_teach": ["PE"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T026", "name": "Nisha Kapoor", "subjects_can_teach": ["PE"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T027", "name": "Vikram Chauhan", "subjects_can_teach": ["PE"], "min_periods_day": 0, "max_periods_day": 8, "min_periods_week": 0, "max_periods_week": 40, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    
    # Computer Teachers (3 teachers for 10 classes)
    {"teacher_id": "T028", "name": "Neha Gupta", "subjects_can_teach": ["COMP"], "min_periods_day": 0, "max_periods_day": 6, "min_periods_week": 0, "max_periods_week": 35, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T029", "name": "Manoj Kumar", "subjects_can_teach": ["COMP"], "min_periods_day": 0, "max_periods_day": 6, "min_periods_week": 0, "max_periods_week": 35, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T030", "name": "Swati Joshi", "subjects_can_teach": ["COMP"], "min_periods_day": 0, "max_periods_day": 6, "min_periods_week": 0, "max_periods_week": 35, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    
    # Art Teachers (2 teachers for 10 classes)
    {"teacher_id": "T031", "name": "Kiran Das", "subjects_can_teach": ["ART"], "min_periods_day": 0, "max_periods_day": 6, "min_periods_week": 0, "max_periods_week": 30, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
    {"teacher_id": "T032", "name": "Pooja Mathur", "subjects_can_teach": ["ART"], "min_periods_day": 0, "max_periods_day": 6, "min_periods_week": 0, "max_periods_week": 30, "max_consecutive_periods": 3, "availability_days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], "availability_time": "08:00-15:30"},
]

SAMPLE_SUBJECTS = [
    {"subject_code": "MATH", "name": "Mathematics", "category": "core", "min_weekly": 6, "max_weekly": 7, "block_required": False, "block_length": 0, "resource_type": None},
    {"subject_code": "SCI", "name": "Science", "category": "core", "min_weekly": 5, "max_weekly": 6, "block_required": False, "block_length": 0, "resource_type": None},
    {"subject_code": "ENG", "name": "English", "category": "language", "min_weekly": 5, "max_weekly": 6, "block_required": False, "block_length": 0, "resource_type": None},
    {"subject_code": "SST", "name": "Social Studies", "category": "core", "min_weekly": 4, "max_weekly": 5, "block_required": False, "block_length": 0, "resource_type": None},
    {"subject_code": "HINDI", "name": "Hindi", "category": "language", "min_weekly": 4, "max_weekly": 5, "block_required": False, "block_length": 0, "resource_type": None},
    {"subject_code": "KANNADA", "name": "Kannada", "category": "language", "min_weekly": 4, "max_weekly": 5, "block_required": False, "block_length": 0, "resource_type": None},
    {"subject_code": "SANSKRIT", "name": "Sanskrit", "category": "language", "min_weekly": 4, "max_weekly": 5, "block_required": False, "block_length": 0, "resource_type": None},
    {"subject_code": "PHY_LAB", "name": "Physics Lab", "category": "lab", "min_weekly": 2, "max_weekly": 2, "block_required": True, "block_length": 2, "resource_type": "Physics Lab"},
    {"subject_code": "CHEM_LAB", "name": "Chemistry Lab", "category": "lab", "min_weekly": 2, "max_weekly": 2, "block_required": True, "block_length": 2, "resource_type": "Chemistry Lab"},
    {"subject_code": "COMP", "name": "Computer Science", "category": "leisure", "min_weekly": 2, "max_weekly": 3, "block_required": False, "block_length": 0, "resource_type": "Computer Lab"},
    {"subject_code": "PE", "name": "Physical Education", "category": "leisure", "min_weekly": 2, "max_weekly": 2, "block_required": False, "block_length": 0, "resource_type": "Sports Ground"},
    {"subject_code": "ART", "name": "Art & Craft", "category": "leisure", "min_weekly": 1, "max_weekly": 2, "block_required": False, "block_length": 0, "resource_type": None},
]

SAMPLE_CLASSES = [
    {"section_id": "6A", "grade": 6, "section": "A", "student_count": 35, "class_teacher_id": "T001", "language_block_enabled": True, "language_structure": "2nd_lang_fixed:ENGLISH"},
    {"section_id": "6B", "grade": 6, "section": "B", "student_count": 38, "class_teacher_id": "T002", "language_block_enabled": True, "language_structure": "2nd_lang_fixed:ENGLISH"},
    {"section_id": "7A", "grade": 7, "section": "A", "student_count": 40, "class_teacher_id": "T005", "language_block_enabled": True, "language_structure": "2nd_lang_fixed:ENGLISH"},
    {"section_id": "7B", "grade": 7, "section": "B", "student_count": 37, "class_teacher_id": "T006", "language_block_enabled": True, "language_structure": "2nd_lang_fixed:ENGLISH"},
    {"section_id": "8A", "grade": 8, "section": "A", "student_count": 42, "class_teacher_id": "T009", "language_block_enabled": True, "language_structure": "2nd_lang_fixed:ENGLISH"},
    {"section_id": "8B", "grade": 8, "section": "B", "student_count": 39, "class_teacher_id": "T010", "language_block_enabled": True, "language_structure": "2nd_lang_fixed:ENGLISH"},
    {"section_id": "9A", "grade": 9, "section": "A", "student_count": 35, "class_teacher_id": "T013", "language_block_enabled": True, "language_structure": "2nd_lang_fixed:ENGLISH"},
    {"section_id": "9B", "grade": 9, "section": "B", "student_count": 33, "class_teacher_id": "T014", "language_block_enabled": True, "language_structure": "2nd_lang_fixed:ENGLISH"},
    {"section_id": "10A", "grade": 10, "section": "A", "student_count": 30, "class_teacher_id": "T003", "language_block_enabled": True, "language_structure": "2nd_lang_fixed:ENGLISH"},
    {"section_id": "10B", "grade": 10, "section": "B", "student_count": 28, "class_teacher_id": "T004", "language_block_enabled": True, "language_structure": "2nd_lang_fixed:ENGLISH"},
]

SAMPLE_MAPPINGS = [
    # 6A - Set 1: Hindi T016, Kannada T019, Sanskrit T022
    {"section_id": "6A", "subject_code": "MATH", "teacher_id": "T001", "is_class_teacher": True},
    {"section_id": "6A", "subject_code": "SCI", "teacher_id": "T005", "is_class_teacher": False},
    {"section_id": "6A", "subject_code": "ENG", "teacher_id": "T009", "is_class_teacher": False},
    {"section_id": "6A", "subject_code": "SST", "teacher_id": "T013", "is_class_teacher": False},
    {"section_id": "6A", "subject_code": "HINDI", "teacher_id": "T016", "is_class_teacher": False},
    {"section_id": "6A", "subject_code": "KANNADA", "teacher_id": "T019", "is_class_teacher": False},
    {"section_id": "6A", "subject_code": "SANSKRIT", "teacher_id": "T022", "is_class_teacher": False},
    {"section_id": "6A", "subject_code": "PE", "teacher_id": "T025", "is_class_teacher": False},
    {"section_id": "6A", "subject_code": "COMP", "teacher_id": "T028", "is_class_teacher": False},
    {"section_id": "6A", "subject_code": "ART", "teacher_id": "T031", "is_class_teacher": False},
    
    # 6B - Set 1: Hindi T016, Kannada T019, Sanskrit T022
    {"section_id": "6B", "subject_code": "MATH", "teacher_id": "T002", "is_class_teacher": True},
    {"section_id": "6B", "subject_code": "SCI", "teacher_id": "T006", "is_class_teacher": False},
    {"section_id": "6B", "subject_code": "ENG", "teacher_id": "T010", "is_class_teacher": False},
    {"section_id": "6B", "subject_code": "SST", "teacher_id": "T014", "is_class_teacher": False},
    {"section_id": "6B", "subject_code": "HINDI", "teacher_id": "T016", "is_class_teacher": False},
    {"section_id": "6B", "subject_code": "KANNADA", "teacher_id": "T019", "is_class_teacher": False},
    {"section_id": "6B", "subject_code": "SANSKRIT", "teacher_id": "T022", "is_class_teacher": False},
    {"section_id": "6B", "subject_code": "PE", "teacher_id": "T026", "is_class_teacher": False},
    {"section_id": "6B", "subject_code": "COMP", "teacher_id": "T029", "is_class_teacher": False},
    {"section_id": "6B", "subject_code": "ART", "teacher_id": "T032", "is_class_teacher": False},
    
    # 7A - Set 2: Hindi T017, Kannada T020, Sanskrit T023
    {"section_id": "7A", "subject_code": "MATH", "teacher_id": "T001", "is_class_teacher": False},
    {"section_id": "7A", "subject_code": "SCI", "teacher_id": "T005", "is_class_teacher": True},
    {"section_id": "7A", "subject_code": "ENG", "teacher_id": "T009", "is_class_teacher": False},
    {"section_id": "7A", "subject_code": "SST", "teacher_id": "T013", "is_class_teacher": False},
    {"section_id": "7A", "subject_code": "HINDI", "teacher_id": "T017", "is_class_teacher": False},
    {"section_id": "7A", "subject_code": "KANNADA", "teacher_id": "T020", "is_class_teacher": False},
    {"section_id": "7A", "subject_code": "SANSKRIT", "teacher_id": "T023", "is_class_teacher": False},
    {"section_id": "7A", "subject_code": "PE", "teacher_id": "T025", "is_class_teacher": False},
    {"section_id": "7A", "subject_code": "COMP", "teacher_id": "T028", "is_class_teacher": False},
    {"section_id": "7A", "subject_code": "ART", "teacher_id": "T031", "is_class_teacher": False},
    
    # 7B - Set 2: Hindi T017, Kannada T020, Sanskrit T023
    {"section_id": "7B", "subject_code": "MATH", "teacher_id": "T002", "is_class_teacher": False},
    {"section_id": "7B", "subject_code": "SCI", "teacher_id": "T006", "is_class_teacher": True},
    {"section_id": "7B", "subject_code": "ENG", "teacher_id": "T010", "is_class_teacher": False},
    {"section_id": "7B", "subject_code": "SST", "teacher_id": "T014", "is_class_teacher": False},
    {"section_id": "7B", "subject_code": "HINDI", "teacher_id": "T017", "is_class_teacher": False},
    {"section_id": "7B", "subject_code": "KANNADA", "teacher_id": "T020", "is_class_teacher": False},
    {"section_id": "7B", "subject_code": "SANSKRIT", "teacher_id": "T023", "is_class_teacher": False},
    {"section_id": "7B", "subject_code": "PE", "teacher_id": "T026", "is_class_teacher": False},
    {"section_id": "7B", "subject_code": "COMP", "teacher_id": "T029", "is_class_teacher": False},
    {"section_id": "7B", "subject_code": "ART", "teacher_id": "T032", "is_class_teacher": False},
    
    # 8A - Set 3: Hindi T018, Kannada T021, Sanskrit T024
    {"section_id": "8A", "subject_code": "MATH", "teacher_id": "T003", "is_class_teacher": False},
    {"section_id": "8A", "subject_code": "SCI", "teacher_id": "T007", "is_class_teacher": False},
    {"section_id": "8A", "subject_code": "ENG", "teacher_id": "T009", "is_class_teacher": True},
    {"section_id": "8A", "subject_code": "SST", "teacher_id": "T015", "is_class_teacher": False},
    {"section_id": "8A", "subject_code": "PHY_LAB", "teacher_id": "T005", "is_class_teacher": False},
    {"section_id": "8A", "subject_code": "CHEM_LAB", "teacher_id": "T007", "is_class_teacher": False},
    {"section_id": "8A", "subject_code": "HINDI", "teacher_id": "T018", "is_class_teacher": False},
    {"section_id": "8A", "subject_code": "KANNADA", "teacher_id": "T021", "is_class_teacher": False},
    {"section_id": "8A", "subject_code": "SANSKRIT", "teacher_id": "T024", "is_class_teacher": False},
    {"section_id": "8A", "subject_code": "PE", "teacher_id": "T027", "is_class_teacher": False},
    {"section_id": "8A", "subject_code": "COMP", "teacher_id": "T030", "is_class_teacher": False},
    
    # 8B - Set 3: Hindi T018, Kannada T021, Sanskrit T024
    {"section_id": "8B", "subject_code": "MATH", "teacher_id": "T004", "is_class_teacher": False},
    {"section_id": "8B", "subject_code": "SCI", "teacher_id": "T008", "is_class_teacher": False},
    {"section_id": "8B", "subject_code": "ENG", "teacher_id": "T010", "is_class_teacher": True},
    {"section_id": "8B", "subject_code": "SST", "teacher_id": "T013", "is_class_teacher": False},
    {"section_id": "8B", "subject_code": "PHY_LAB", "teacher_id": "T006", "is_class_teacher": False},
    {"section_id": "8B", "subject_code": "CHEM_LAB", "teacher_id": "T008", "is_class_teacher": False},
    {"section_id": "8B", "subject_code": "HINDI", "teacher_id": "T018", "is_class_teacher": False},
    {"section_id": "8B", "subject_code": "KANNADA", "teacher_id": "T021", "is_class_teacher": False},
    {"section_id": "8B", "subject_code": "SANSKRIT", "teacher_id": "T024", "is_class_teacher": False},
    {"section_id": "8B", "subject_code": "PE", "teacher_id": "T025", "is_class_teacher": False},
    {"section_id": "8B", "subject_code": "COMP", "teacher_id": "T028", "is_class_teacher": False},
    
    # 9A - Set 4: Hindi T033, Kannada T034, Sanskrit T035
    {"section_id": "9A", "subject_code": "MATH", "teacher_id": "T001", "is_class_teacher": False},
    {"section_id": "9A", "subject_code": "SCI", "teacher_id": "T005", "is_class_teacher": False},
    {"section_id": "9A", "subject_code": "ENG", "teacher_id": "T011", "is_class_teacher": False},
    {"section_id": "9A", "subject_code": "SST", "teacher_id": "T013", "is_class_teacher": True},
    {"section_id": "9A", "subject_code": "PHY_LAB", "teacher_id": "T005", "is_class_teacher": False},
    {"section_id": "9A", "subject_code": "CHEM_LAB", "teacher_id": "T007", "is_class_teacher": False},
    {"section_id": "9A", "subject_code": "HINDI", "teacher_id": "T033", "is_class_teacher": False},
    {"section_id": "9A", "subject_code": "KANNADA", "teacher_id": "T034", "is_class_teacher": False},
    {"section_id": "9A", "subject_code": "SANSKRIT", "teacher_id": "T035", "is_class_teacher": False},
    {"section_id": "9A", "subject_code": "PE", "teacher_id": "T026", "is_class_teacher": False},
    {"section_id": "9A", "subject_code": "COMP", "teacher_id": "T029", "is_class_teacher": False},
    
    # 9B - Set 4: Hindi T033, Kannada T034, Sanskrit T035
    {"section_id": "9B", "subject_code": "MATH", "teacher_id": "T002", "is_class_teacher": False},
    {"section_id": "9B", "subject_code": "SCI", "teacher_id": "T006", "is_class_teacher": False},
    {"section_id": "9B", "subject_code": "ENG", "teacher_id": "T012", "is_class_teacher": False},
    {"section_id": "9B", "subject_code": "SST", "teacher_id": "T014", "is_class_teacher": True},
    {"section_id": "9B", "subject_code": "PHY_LAB", "teacher_id": "T006", "is_class_teacher": False},
    {"section_id": "9B", "subject_code": "CHEM_LAB", "teacher_id": "T008", "is_class_teacher": False},
    {"section_id": "9B", "subject_code": "HINDI", "teacher_id": "T033", "is_class_teacher": False},
    {"section_id": "9B", "subject_code": "KANNADA", "teacher_id": "T034", "is_class_teacher": False},
    {"section_id": "9B", "subject_code": "SANSKRIT", "teacher_id": "T035", "is_class_teacher": False},
    {"section_id": "9B", "subject_code": "PE", "teacher_id": "T027", "is_class_teacher": False},
    {"section_id": "9B", "subject_code": "COMP", "teacher_id": "T030", "is_class_teacher": False},
    
    # 10A - Set 4: Hindi T033, Kannada T034, Sanskrit T035
    {"section_id": "10A", "subject_code": "MATH", "teacher_id": "T003", "is_class_teacher": True},
    {"section_id": "10A", "subject_code": "SCI", "teacher_id": "T007", "is_class_teacher": False},
    {"section_id": "10A", "subject_code": "ENG", "teacher_id": "T011", "is_class_teacher": False},
    {"section_id": "10A", "subject_code": "SST", "teacher_id": "T015", "is_class_teacher": False},
    {"section_id": "10A", "subject_code": "PHY_LAB", "teacher_id": "T005", "is_class_teacher": False},
    {"section_id": "10A", "subject_code": "CHEM_LAB", "teacher_id": "T007", "is_class_teacher": False},
    {"section_id": "10A", "subject_code": "HINDI", "teacher_id": "T033", "is_class_teacher": False},
    {"section_id": "10A", "subject_code": "KANNADA", "teacher_id": "T034", "is_class_teacher": False},
    {"section_id": "10A", "subject_code": "SANSKRIT", "teacher_id": "T035", "is_class_teacher": False},
    {"section_id": "10A", "subject_code": "PE", "teacher_id": "T025", "is_class_teacher": False},
    {"section_id": "10A", "subject_code": "COMP", "teacher_id": "T028", "is_class_teacher": False},
    
    # 10B - Set 4: Hindi T033, Kannada T034, Sanskrit T035
    {"section_id": "10B", "subject_code": "MATH", "teacher_id": "T004", "is_class_teacher": True},
    {"section_id": "10B", "subject_code": "SCI", "teacher_id": "T008", "is_class_teacher": False},
    {"section_id": "10B", "subject_code": "ENG", "teacher_id": "T012", "is_class_teacher": False},
    {"section_id": "10B", "subject_code": "SST", "teacher_id": "T015", "is_class_teacher": False},
    {"section_id": "10B", "subject_code": "PHY_LAB", "teacher_id": "T006", "is_class_teacher": False},
    {"section_id": "10B", "subject_code": "CHEM_LAB", "teacher_id": "T008", "is_class_teacher": False},
    {"section_id": "10B", "subject_code": "HINDI", "teacher_id": "T033", "is_class_teacher": False},
    {"section_id": "10B", "subject_code": "KANNADA", "teacher_id": "T034", "is_class_teacher": False},
    {"section_id": "10B", "subject_code": "SANSKRIT", "teacher_id": "T035", "is_class_teacher": False},
    {"section_id": "10B", "subject_code": "PE", "teacher_id": "T026", "is_class_teacher": False},
    {"section_id": "10B", "subject_code": "COMP", "teacher_id": "T029", "is_class_teacher": False},
]

SAMPLE_RESOURCES = [
    {"resource_id": "PHYS_LAB_1", "resource_type": "Physics Lab", "max_simultaneous_capacity": 2, "accessible_grades": [8, 9, 10]},
    {"resource_id": "CHEM_LAB_1", "resource_type": "Chemistry Lab", "max_simultaneous_capacity": 2, "accessible_grades": [8, 9, 10]},
    {"resource_id": "COMP_LAB_1", "resource_type": "Computer Lab", "max_simultaneous_capacity": 3, "accessible_grades": [6, 7, 8, 9, 10]},
    {"resource_id": "SPORTS_GROUND", "resource_type": "Sports Ground", "max_simultaneous_capacity": 2, "accessible_grades": [6, 7, 8, 9, 10]},
]

SAMPLE_LANGUAGE_GROUPS = [
    # Set 1: T016 (Hindi), T019 (Kannada), T022 (Sanskrit) - serves 6A, 6B (only 2 sections)
    {"section_id": "6A", "language_slot": "1st_lang", "hindi_teacher": "T016", "kannada_teacher": "T019", "sanskrit_teacher": "T022"},
    {"section_id": "6B", "language_slot": "1st_lang", "hindi_teacher": "T016", "kannada_teacher": "T019", "sanskrit_teacher": "T022"},
    
    # Set 2: T017 (Hindi), T020 (Kannada), T023 (Sanskrit) - serves 7A, 7B (only 2 sections)
    {"section_id": "7A", "language_slot": "1st_lang", "hindi_teacher": "T017", "kannada_teacher": "T020", "sanskrit_teacher": "T023"},
    {"section_id": "7B", "language_slot": "1st_lang", "hindi_teacher": "T017", "kannada_teacher": "T020", "sanskrit_teacher": "T023"},
    
    # Set 3: T018 (Hindi), T021 (Kannada), T024 (Sanskrit) - serves 8A, 8B (only 2 sections)
    {"section_id": "8A", "language_slot": "1st_lang", "hindi_teacher": "T018", "kannada_teacher": "T021", "sanskrit_teacher": "T024"},
    {"section_id": "8B", "language_slot": "1st_lang", "hindi_teacher": "T018", "kannada_teacher": "T021", "sanskrit_teacher": "T024"},
    
    # Set 4: T033 (Hindi), T034 (Kannada), T035 (Sanskrit) - serves 9A, 9B, 10A, 10B (4 sections)
    {"section_id": "9A", "language_slot": "1st_lang", "hindi_teacher": "T033", "kannada_teacher": "T034", "sanskrit_teacher": "T035"},
    {"section_id": "9B", "language_slot": "1st_lang", "hindi_teacher": "T033", "kannada_teacher": "T034", "sanskrit_teacher": "T035"},
    {"section_id": "10A", "language_slot": "1st_lang", "hindi_teacher": "T033", "kannada_teacher": "T034", "sanskrit_teacher": "T035"},
    {"section_id": "10B", "language_slot": "1st_lang", "hindi_teacher": "T033", "kannada_teacher": "T034", "sanskrit_teacher": "T035"},
]

SAMPLE_CONSTRAINTS_CONFIG = {
    "prayer_enabled": True,
    "language_sync_enabled": False,  # DISABLED by default - strict sync is too constraining
    "class_teacher_period_1": True,  # ENABLED - Class teacher must take period 1
    "no_subject_twice_daily": False, # DISABLED by default - allows Math twice if needed
    "substitution_reserve_count": 0, # DISABLED - no reserve requirement by default
    "core_morning_only": False,      # DISABLED by default - soft constraint handles this
    "max_consecutive_default": 3,
    "max_daily_load_variance": 3,
    "soft_weight_core_morning": 10,
    "soft_weight_teacher_balance": 8,
    "soft_weight_minimize_gaps": 5,
    "soft_weight_leisure_afternoon": 3,
    "soft_weight_avoid_pe_period_1": 4,
    "soft_weight_subject_distribution": 3,
    "soft_weight_teacher_free_period": 2,
    "soft_weight_fair_slot_distribution": 5,
}


def _generate_csv_content():
    """Generate multi-sheet CSV content as a bundle."""
    import csv
    from io import StringIO
    
    sheets = {}
    
    # School Config
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['school_id', 'name', 'academic_year', 'start_time', 'end_time', 'weekdays', 
                     'saturday_periods', 'prayer_minutes', 'lunch_period_index', 'recess_period_indices'])
    writer.writerow([
        SAMPLE_SCHOOL['school_id'], SAMPLE_SCHOOL['name'], SAMPLE_SCHOOL['academic_year'],
        SAMPLE_SCHOOL['start_time'], SAMPLE_SCHOOL['end_time'],
        ','.join(SAMPLE_SCHOOL['weekdays']), SAMPLE_SCHOOL['saturday_periods'],
        SAMPLE_SCHOOL['prayer_duration_minutes'], SAMPLE_SCHOOL['lunch_period_index'],
        ','.join(map(str, SAMPLE_SCHOOL['recess_period_indices']))
    ])
    sheets['school_config'] = output.getvalue()
    
    # Periods
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['period_index', 'start_time', 'end_time', 'duration_minutes', 'is_prayer', 'is_recess', 'is_lunch'])
    for p in SAMPLE_PERIODS:
        writer.writerow([p['period_index'], p['start_time'], p['end_time'], p['duration_minutes'], 
                        1 if p['is_prayer'] else 0, 1 if p['is_recess'] else 0, 1 if p['is_lunch'] else 0])
    sheets['periods'] = output.getvalue()
    
    # Classes
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['section_id', 'grade', 'section', 'student_count', 'class_teacher_id', 
                     'language_block_enabled', 'language_structure'])
    for c in SAMPLE_CLASSES:
        writer.writerow([c['section_id'], c['grade'], c['section'], c['student_count'], 
                        c['class_teacher_id'], 1 if c['language_block_enabled'] else 0, 
                        c['language_structure'] or ''])
    sheets['classes'] = output.getvalue()
    
    # Subjects
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['subject_code', 'name', 'category', 'min_weekly', 'max_weekly', 
                     'block_required', 'block_length', 'resource_type'])
    for s in SAMPLE_SUBJECTS:
        writer.writerow([s['subject_code'], s['name'], s['category'], s['min_weekly'], 
                        s['max_weekly'], 1 if s['block_required'] else 0, s['block_length'], 
                        s['resource_type'] or ''])
    sheets['subjects'] = output.getvalue()
    
    # Teachers
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['teacher_id', 'name', 'subjects_can_teach', 'min_periods_day', 'max_periods_day',
                     'min_periods_week', 'max_periods_week', 'max_consecutive_periods', 
                     'availability_days', 'availability_time'])
    for t in SAMPLE_TEACHERS:
        writer.writerow([t['teacher_id'], t['name'], ','.join(t['subjects_can_teach']),
                        t['min_periods_day'], t['max_periods_day'], t['min_periods_week'],
                        t['max_periods_week'], t['max_consecutive_periods'],
                        ','.join(t['availability_days']), t['availability_time']])
    sheets['teachers'] = output.getvalue()
    
    # Mappings
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['section_id', 'subject_code', 'teacher_id', 'is_class_teacher'])
    for m in SAMPLE_MAPPINGS:
        writer.writerow([m['section_id'], m['subject_code'], m['teacher_id'], 
                        1 if m['is_class_teacher'] else 0])
    sheets['mappings'] = output.getvalue()
    
    # Resources
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['resource_id', 'resource_type', 'max_simultaneous_capacity', 'accessible_grades'])
    for r in SAMPLE_RESOURCES:
        writer.writerow([r['resource_id'], r['resource_type'], r['max_simultaneous_capacity'],
                        ','.join(map(str, r['accessible_grades']))])
    sheets['resources'] = output.getvalue()
    
    # Language Groups
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['section_id', 'language_slot', 'hindi_teacher', 'kannada_teacher', 'sanskrit_teacher'])
    for lg in SAMPLE_LANGUAGE_GROUPS:
        writer.writerow([lg['section_id'], lg['language_slot'], lg['hindi_teacher'],
                        lg['kannada_teacher'], lg['sanskrit_teacher']])
    sheets['language_groups'] = output.getvalue()
    
    # Constraints Config
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['constraint_name', 'enabled', 'value'])
    for name, value in SAMPLE_CONSTRAINTS_CONFIG.items():
        if isinstance(value, bool):
            writer.writerow([name, 1 if value else 0, ''])
        else:
            writer.writerow([name, 1, value])
    sheets['constraints_config'] = output.getvalue()
    
    return sheets


@router.get("/download-template")
async def download_template():
    """
    Download an empty Excel template with all required sheets and headers.
    
    Returns:
        StreamingResponse with Excel file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV format
        sheets = _generate_csv_content()
        # Return as JSON for now if openpyxl not available
        return JSONResponse({
            "message": "openpyxl not installed, returning CSV templates",
            "sheets": {name: content for name, content in sheets.items()}
        })
    
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet definitions
    sheet_definitions = {
        "school_config": {
            "headers": ["school_id", "name", "academic_year", "start_time", "end_time", "weekdays",
                       "saturday_periods", "prayer_minutes", "lunch_period_index", "recess_period_indices"],
            "widths": [10, 30, 15, 12, 12, 30, 18, 15, 20, 22],
            "example": [1, "Your School Name", "2025-2026", "08:00", "15:30", "Mon,Tue,Wed,Thu,Fri,Sat", 4, 30, 5, "3"]
        },
        "periods": {
            "headers": ["period_index", "start_time", "end_time", "duration_minutes", "is_prayer", "is_recess", "is_lunch"],
            "widths": [12, 12, 12, 18, 10, 10, 10],
            "example": [0, "08:00", "08:30", 30, 1, 0, 0]
        },
        "classes": {
            "headers": ["section_id", "grade", "section", "student_count", "class_teacher_id", 
                       "language_block_enabled", "language_structure"],
            "widths": [12, 8, 10, 14, 18, 22, 25],
            "example": ["8A", 8, "A", 40, "T001", 1, "2nd_lang_fixed:ENGLISH"]
        },
        "subjects": {
            "headers": ["subject_code", "name", "category", "min_weekly", "max_weekly", 
                       "block_required", "block_length", "resource_type"],
            "widths": [15, 25, 12, 12, 12, 15, 12, 20],
            "example": ["MATH", "Mathematics", "core", 6, 7, 0, 0, ""]
        },
        "teachers": {
            "headers": ["teacher_id", "name", "subjects_can_teach", "min_periods_day", "max_periods_day",
                       "min_periods_week", "max_periods_week", "max_consecutive_periods", 
                       "availability_days", "availability_time"],
            "widths": [12, 20, 25, 16, 16, 18, 18, 22, 30, 18],
            "example": ["T001", "Teacher Name", "MATH,SCI", 3, 6, 18, 30, 3, "Mon,Tue,Wed,Thu,Fri,Sat", "08:00-15:30"]
        },
        "mappings": {
            "headers": ["section_id", "subject_code", "teacher_id", "is_class_teacher"],
            "widths": [12, 15, 12, 16],
            "example": ["8A", "MATH", "T001", 1]
        },
        "resources": {
            "headers": ["resource_id", "resource_type", "max_simultaneous_capacity", "accessible_grades"],
            "widths": [15, 20, 26, 20],
            "example": ["COMP_LAB_1", "Computer Lab", 3, "6,7,8,9,10"]
        },
        "language_groups": {
            "headers": ["section_id", "language_slot", "hindi_teacher", "kannada_teacher", "sanskrit_teacher"],
            "widths": [12, 15, 16, 18, 18],
            "example": ["8A", "1st_lang", "T011", "T012", "T013"]
        },
        "constraints_config": {
            "headers": ["constraint_name", "enabled", "value"],
            "widths": [35, 10, 10],
            "example": ["prayer_enabled", 1, ""]
        },
    }
    
    # Remove default sheet
    wb.remove(wb.active)
    
    for sheet_name, definition in sheet_definitions.items():
        ws = wb.create_sheet(title=sheet_name)
        
        # Add headers
        for col, header in enumerate(definition["headers"], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        for col, width in enumerate(definition["widths"], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Add example row (commented)
        for col, value in enumerate(definition["example"], 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = border
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=timetable_template.xlsx"
        }
    )


@router.get("/download-sample")
async def download_sample():
    """
    Download the complete Vidya Mandir sample dataset as Excel.
    
    Returns:
        StreamingResponse with Excel file containing all sample data
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to JSON format
        return JSONResponse({
            "school": SAMPLE_SCHOOL,
            "periods": SAMPLE_PERIODS,
            "teachers": SAMPLE_TEACHERS,
            "subjects": SAMPLE_SUBJECTS,
            "classes": SAMPLE_CLASSES,
            "mappings": SAMPLE_MAPPINGS,
            "resources": SAMPLE_RESOURCES,
            "language_groups": SAMPLE_LANGUAGE_GROUPS,
            "constraints_config": SAMPLE_CONSTRAINTS_CONFIG,
        })
    
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def add_sheet(name, headers, data, widths):
        ws = wb.create_sheet(title=name)
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Add data
        for row_idx, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # School Config
    add_sheet("school_config",
              ["school_id", "name", "academic_year", "start_time", "end_time", "weekdays",
               "saturday_periods", "prayer_minutes", "lunch_period_index", "recess_period_indices"],
              [[SAMPLE_SCHOOL['school_id'], SAMPLE_SCHOOL['name'], SAMPLE_SCHOOL['academic_year'],
                SAMPLE_SCHOOL['start_time'], SAMPLE_SCHOOL['end_time'],
                ','.join(SAMPLE_SCHOOL['weekdays']), SAMPLE_SCHOOL['saturday_periods'],
                SAMPLE_SCHOOL['prayer_duration_minutes'], SAMPLE_SCHOOL['lunch_period_index'],
                ','.join(map(str, SAMPLE_SCHOOL['recess_period_indices']))]],
              [10, 30, 15, 12, 12, 30, 18, 15, 20, 22])
    
    # Periods
    add_sheet("periods",
              ["period_index", "start_time", "end_time", "duration_minutes", "is_prayer", "is_recess", "is_lunch"],
              [[p['period_index'], p['start_time'], p['end_time'], p['duration_minutes'],
                1 if p['is_prayer'] else 0, 1 if p['is_recess'] else 0, 1 if p['is_lunch'] else 0]
               for p in SAMPLE_PERIODS],
              [12, 12, 12, 18, 10, 10, 10])
    
    # Classes
    add_sheet("classes",
              ["section_id", "grade", "section", "student_count", "class_teacher_id",
               "language_block_enabled", "language_structure"],
              [[c['section_id'], c['grade'], c['section'], c['student_count'],
                c['class_teacher_id'], 1 if c['language_block_enabled'] else 0,
                c['language_structure'] or ''] for c in SAMPLE_CLASSES],
              [12, 8, 10, 14, 18, 22, 25])
    
    # Subjects
    add_sheet("subjects",
              ["subject_code", "name", "category", "min_weekly", "max_weekly",
               "block_required", "block_length", "resource_type"],
              [[s['subject_code'], s['name'], s['category'], s['min_weekly'],
                s['max_weekly'], 1 if s['block_required'] else 0, s['block_length'],
                s['resource_type'] or ''] for s in SAMPLE_SUBJECTS],
              [15, 25, 12, 12, 12, 15, 12, 20])
    
    # Teachers
    add_sheet("teachers",
              ["teacher_id", "name", "subjects_can_teach", "min_periods_day", "max_periods_day",
               "min_periods_week", "max_periods_week", "max_consecutive_periods",
               "availability_days", "availability_time"],
              [[t['teacher_id'], t['name'], ','.join(t['subjects_can_teach']),
                t['min_periods_day'], t['max_periods_day'], t['min_periods_week'],
                t['max_periods_week'], t['max_consecutive_periods'],
                ','.join(t['availability_days']), t['availability_time']]
               for t in SAMPLE_TEACHERS],
              [12, 20, 25, 16, 16, 18, 18, 22, 30, 18])
    
    # Mappings
    add_sheet("mappings",
              ["section_id", "subject_code", "teacher_id", "is_class_teacher"],
              [[m['section_id'], m['subject_code'], m['teacher_id'],
                1 if m['is_class_teacher'] else 0] for m in SAMPLE_MAPPINGS],
              [12, 15, 12, 16])
    
    # Resources
    add_sheet("resources",
              ["resource_id", "resource_type", "max_simultaneous_capacity", "accessible_grades"],
              [[r['resource_id'], r['resource_type'], r['max_simultaneous_capacity'],
                ','.join(map(str, r['accessible_grades']))] for r in SAMPLE_RESOURCES],
              [15, 20, 26, 20])
    
    # Language Groups
    add_sheet("language_groups",
              ["section_id", "language_slot", "hindi_teacher", "kannada_teacher", "sanskrit_teacher"],
              [[lg['section_id'], lg['language_slot'], lg['hindi_teacher'],
                lg['kannada_teacher'], lg['sanskrit_teacher']] for lg in SAMPLE_LANGUAGE_GROUPS],
              [12, 15, 16, 18, 18])
    
    # Constraints Config
    constraints_data = []
    for name, value in SAMPLE_CONSTRAINTS_CONFIG.items():
        if isinstance(value, bool):
            constraints_data.append([name, 1 if value else 0, ''])
        else:
            constraints_data.append([name, 1, value])
    add_sheet("constraints_config",
              ["constraint_name", "enabled", "value"],
              constraints_data,
              [35, 10, 10])
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=vidya_mandir_sample.xlsx"
        }
    )


@router.get("/sample-data")
async def get_sample_data():
    """
    Get sample data as JSON for auto-loading into the frontend.
    
    Returns:
        JSON with all sample data transformed for frontend consumption
    """
    # Transform to frontend-compatible format
    teachers = []
    for t in SAMPLE_TEACHERS:
        teachers.append({
            "teacher_id": t["teacher_id"],
            "name": t["name"],
            "subjects_can_teach": t["subjects_can_teach"],
            "sections_assigned": [],
            "min_periods_day": t["min_periods_day"],
            "max_periods_day": t["max_periods_day"],
            "min_periods_week": t["min_periods_week"],
            "max_periods_week": t["max_periods_week"],
            "max_consecutive_periods": t["max_consecutive_periods"],
            "is_class_teacher_of": None,
            "is_specialist": len(t["availability_days"]) < 6,
            "availability": {
                day: {
                    "available": day in t["availability_days"],
                    "from_time": t["availability_time"].split("-")[0] if day in t["availability_days"] else None,
                    "to_time": t["availability_time"].split("-")[1] if day in t["availability_days"] else None,
                }
                for day in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
            }
        })
    
    # Update class teacher info
    for c in SAMPLE_CLASSES:
        for t in teachers:
            if t["teacher_id"] == c["class_teacher_id"]:
                t["is_class_teacher_of"] = c["section_id"]
                break
    
    # Update sections assigned based on mappings
    for m in SAMPLE_MAPPINGS:
        for t in teachers:
            if t["teacher_id"] == m["teacher_id"]:
                if m["section_id"] not in t["sections_assigned"]:
                    t["sections_assigned"].append(m["section_id"])
                break
    
    subjects = []
    for s in SAMPLE_SUBJECTS:
        subjects.append({
            "subject_id": s["subject_code"],
            "name": s["name"],
            "category": s["category"],
            "min_per_week": s["min_weekly"],
            "max_per_week": s["max_weekly"],
            "requires_block": s["block_required"],
            "block_length": s["block_length"] if s["block_required"] else 2,
            "requires_resource": s["resource_type"] is not None,
            "resource_type": s["resource_type"],
            "prefer_morning": s["category"] == "core",
            "avoid_after_lunch": s["category"] == "core",
            "is_specialist": False,
            "is_language_block": s["category"] == "language" and s["subject_code"] not in ["ENG"],
        })
    
    classes = []
    for c in SAMPLE_CLASSES:
        # Build subject-teacher map
        subject_teacher_map = {}
        language_subjects = []
        language_teachers = []
        
        for m in SAMPLE_MAPPINGS:
            if m["section_id"] == c["section_id"]:
                subject_teacher_map[m["subject_code"]] = m["teacher_id"]
        
        # Get language block info
        for lg in SAMPLE_LANGUAGE_GROUPS:
            if lg["section_id"] == c["section_id"]:
                language_subjects = ["HINDI", "KANNADA", "SANSKRIT"]
                language_teachers = [lg["hindi_teacher"], lg["kannada_teacher"], lg["sanskrit_teacher"]]
                break
        
        classes.append({
            "section_id": c["section_id"],
            "grade": c["grade"],
            "section_name": c["section"],
            "class_teacher_id": c["class_teacher_id"],
            "subject_teacher_map": subject_teacher_map,
            "language_block_enabled": c["language_block_enabled"],
            "language_subjects": language_subjects,
            "language_teachers": language_teachers,
        })
    
    resources = []
    for r in SAMPLE_RESOURCES:
        resources.append({
            "resource_id": r["resource_id"],
            "resource_type": r["resource_type"],
            "name": r["resource_type"],
            "max_simultaneous_capacity": r["max_simultaneous_capacity"],
            "available_periods": None,
        })
    
    school = {
        "school_id": SAMPLE_SCHOOL["school_id"],
        "name": SAMPLE_SCHOOL["name"],
        "start_time": SAMPLE_SCHOOL["start_time"],
        "end_time": SAMPLE_SCHOOL["end_time"],
        "weekdays": SAMPLE_SCHOOL["weekdays"],
        "periods_per_weekday": SAMPLE_SCHOOL["periods_per_weekday"],
        "saturday_periods": SAMPLE_SCHOOL["saturday_periods"],
        "period_duration_minutes": SAMPLE_SCHOOL["period_duration_minutes"],
        "prayer_enabled": SAMPLE_SCHOOL["prayer_enabled"],
        "prayer_duration_minutes": SAMPLE_SCHOOL["prayer_duration_minutes"],
        "lunch_period_index": SAMPLE_SCHOOL["lunch_period_index"],
        "lunch_after_period": SAMPLE_SCHOOL["lunch_after_period"],
        "lunch_duration_minutes": SAMPLE_SCHOOL["lunch_duration_minutes"],
        "recess_period_indices": SAMPLE_SCHOOL["recess_period_indices"],
        "recess_after_every_n_periods": SAMPLE_SCHOOL["recess_after_every_n_periods"],
        "recess_duration_minutes": SAMPLE_SCHOOL["recess_duration_minutes"],
    }
    
    return {
        "upload_id": "sample-data-vidya-mandir",
        "preview": {
            "teachers": len(teachers),
            "classes": len(classes),
            "subjects": len(subjects),
            "resources": len(resources),
        },
        "sample_rows": {
            "teachers": teachers[:3],
            "classes": classes[:3],
            "subjects": subjects[:3],
        },
        "validation_errors": [],
        "school": school,
        "teachers": teachers,
        "subjects": subjects,
        "classes": classes,
        "resources": resources,
        "constraints": SAMPLE_CONSTRAINTS_CONFIG,
    }
