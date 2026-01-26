"""
Excel parser for timetable data files.

Handles parsing of XLSX files using openpyxl.
Supports multiple sheets (one per data type).
"""

import logging
from pathlib import Path
from typing import Union

logger = logging.getLogger(__name__)

# Optional import - will fail gracefully if not installed
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel parsing will not be available.")


def parse_excel(file_path: Union[str, Path], data_type: str) -> list[dict]:
    """
    Parse an Excel file based on its data type.
    
    Can handle:
    1. Single-sheet files where sheet contains data of specified type
    2. Multi-sheet files with sheets named by data type
    
    Args:
        file_path: Path to the Excel file
        data_type: Type of data ('teachers', 'classes', 'subjects', 'resources')
    
    Returns:
        List of parsed dictionaries
    
    Raises:
        ImportError: If openpyxl is not installed
        ValueError: If data_type is not recognized
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel parsing. "
            "Install it with: pip install openpyxl"
        )
    
    file_path = Path(file_path)
    
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    # Try to find the right sheet
    sheet = None
    
    # First, try exact data_type match
    if data_type.lower() in [s.lower() for s in workbook.sheetnames]:
        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() == data_type.lower():
                sheet = workbook[sheet_name]
                break
    
    # Try plural/singular variants
    if sheet is None:
        variants = [data_type, data_type + "s", data_type.rstrip("s")]
        for variant in variants:
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower() == variant.lower():
                    sheet = workbook[sheet_name]
                    break
            if sheet:
                break
    
    # Fall back to first sheet
    if sheet is None:
        sheet = workbook.active
        logger.warning(
            f"No sheet named '{data_type}' found. Using active sheet: {sheet.title}"
        )
    
    # Parse sheet to list of dicts
    rows = list(sheet.iter_rows(values_only=True))
    
    if not rows:
        logger.warning(f"Empty sheet: {sheet.title}")
        return []
    
    # First row is headers
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    
    data = []
    for row in rows[1:]:
        if not any(row):  # Skip empty rows
            continue
        
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                # Convert value to string if not None
                if value is not None:
                    row_dict[headers[i]] = str(value) if not isinstance(value, (int, float, bool)) else value
                else:
                    row_dict[headers[i]] = ""
        
        if row_dict:
            data.append(row_dict)
    
    workbook.close()
    
    # Use CSV parsers to transform the data
    return _transform_excel_data(data, data_type)


def _transform_excel_data(data: list[dict], data_type: str) -> list[dict]:
    """
    Transform raw Excel data using the same logic as CSV parsers.
    
    Args:
        data: List of row dictionaries from Excel
        data_type: Type of data
    
    Returns:
        List of transformed dictionaries
    """
    from .csv_parser import (
        _parse_list,
        _parse_subject_teacher_map,
        _parse_int,
        _parse_bool,
        _normalize_category,
    )
    
    result = []
    
    if data_type.lower() in ("teachers", "teacher"):
        for row in data:
            teacher = {
                "teacher_id": str(row.get("teacher_id", "")).strip(),
                "name": str(row.get("name", "")).strip(),
                "subjects_can_teach": _parse_list(str(row.get("subjects", ""))),
                "min_periods_day": _parse_int(row.get("min_day"), 0),
                "max_periods_day": _parse_int(row.get("max_day"), 8),
            }
            
            if row.get("max_consecutive"):
                teacher["max_consecutive_periods"] = _parse_int(row.get("max_consecutive"), 3)
            if row.get("min_week"):
                teacher["min_periods_week"] = _parse_int(row.get("min_week"), 0)
            if row.get("max_week"):
                teacher["max_periods_week"] = _parse_int(row.get("max_week"), 40)
            if row.get("class_teacher_of"):
                teacher["is_class_teacher_of"] = str(row.get("class_teacher_of")).strip()
            
            if teacher["teacher_id"] and teacher["name"]:
                result.append(teacher)
    
    elif data_type.lower() in ("classes", "class"):
        for row in data:
            cls = {
                "section_id": str(row.get("section_id", "")).strip(),
                "grade": _parse_int(row.get("grade"), 1),
                "subject_teacher_map": _parse_subject_teacher_map(str(row.get("subjects", ""))),
            }
            
            if row.get("section_name"):
                cls["section_name"] = str(row.get("section_name")).strip()
            if row.get("class_teacher_id"):
                cls["class_teacher_id"] = str(row.get("class_teacher_id")).strip()
            if row.get("language_block"):
                cls["language_block_enabled"] = _parse_bool(str(row.get("language_block")))
            
            if cls["section_id"] and cls["subject_teacher_map"]:
                result.append(cls)
    
    elif data_type.lower() in ("subjects", "subject"):
        for row in data:
            subject = {
                "subject_id": str(row.get("subject_id", "")).strip(),
                "name": str(row.get("name", "")).strip(),
                "category": _normalize_category(str(row.get("category", "core"))),
                "min_per_week": _parse_int(row.get("min_week"), 1),
                "max_per_week": _parse_int(row.get("max_week"), 10),
            }
            
            requires_block = _parse_bool(str(row.get("requires_block", "")))
            if requires_block:
                subject["requires_block"] = True
                subject["block_length"] = _parse_int(row.get("block_length"), 2)
            
            resource_type = str(row.get("resource_type", "")).strip()
            if resource_type:
                subject["requires_resource"] = True
                subject["resource_type"] = resource_type
            
            if subject["subject_id"] and subject["name"]:
                result.append(subject)
    
    elif data_type.lower() in ("resources", "resource"):
        for row in data:
            resource = {
                "resource_id": str(row.get("resource_id", "")).strip(),
                "resource_type": str(row.get("resource_type", "")).strip(),
                "max_simultaneous_capacity": _parse_int(row.get("capacity"), 1),
            }
            
            if row.get("name"):
                resource["name"] = str(row.get("name")).strip()
            
            if resource["resource_id"] and resource["resource_type"]:
                result.append(resource)
    
    else:
        raise ValueError(f"Unknown data type: {data_type}")
    
    logger.info(f"Transformed {len(result)} {data_type} records from Excel")
    return result


def parse_excel_workbook(file_path: Union[str, Path]) -> dict[str, list[dict]]:
    """
    Parse all sheets from an Excel workbook.
    
    Returns data from sheets named 'teachers', 'classes', 'subjects', 'resources'.
    
    Args:
        file_path: Path to the Excel file
    
    Returns:
        Dictionary with keys for each data type found
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel parsing")
    
    file_path = Path(file_path)
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    result = {}
    data_types = ["teachers", "classes", "subjects", "resources"]
    
    for data_type in data_types:
        # Check if sheet exists
        sheet_exists = False
        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() == data_type or sheet_name.lower() == data_type.rstrip("s"):
                sheet_exists = True
                break
        
        if sheet_exists:
            try:
                result[data_type] = parse_excel(file_path, data_type)
            except Exception as e:
                logger.error(f"Error parsing {data_type} sheet: {e}")
    
    workbook.close()
    return result
